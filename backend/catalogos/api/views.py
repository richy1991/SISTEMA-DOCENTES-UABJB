import re
import unicodedata

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import AllowAny, BasePermission, SAFE_METHODS, IsAuthenticated
from rest_framework.response import Response
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
import pdfplumber

from catalogos.models import ItemCatalogo, OperacionCatalogo, Direccion
from poa_document.models import UsuarioPOA
from catalogos.api.serializers import (
    ItemCatalogoSerializer,
    OperacionCatalogoSerializer,
    DireccionSerializer,
    PartidaCatalogoSerializer,
)


class IsElaboradorOrReadOnly(BasePermission):
    """Permite lectura a usuarios autenticados y escritura solo a rol elaborador."""

    message = 'Solo usuarios con rol elaborador pueden modificar el catálogo de items.'

    def has_permission(self, request, view):
        user = request.user
        if not user or not user.is_authenticated:
            return False

        if request.method in SAFE_METHODS:
            return True

        return UsuarioPOA.objects.filter(
            user=user,
            activo=True,
            rol='elaborador',
        ).exists()


class ItemCatalogoViewSet(viewsets.ModelViewSet):
    serializer_class = ItemCatalogoSerializer
    permission_classes = [IsElaboradorOrReadOnly]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @staticmethod
    def _normalize_header(value):
        if value is None:
            return ''
        text = str(value).strip().lower()
        text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
        return text.replace(' ', '_')

    @staticmethod
    def _normalize_code(value):
        if value is None:
            return ''
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()

    @action(detail=False, methods=['get', 'post'], url_path='importar-excel')
    def importar_excel(self, request, *args, **kwargs):
        if request.method == 'GET':
            return Response(
                {
                    'detail': 'Endpoint de importación de catálogo.',
                    'uso': 'Envíe POST multipart/form-data con el archivo en el campo "archivo".',
                },
                status=200,
            )

        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'detail': 'Debe adjuntar un archivo Excel en el campo "archivo".'}, status=400)

        dry_run = str(request.data.get('dry_run', 'false')).strip().lower() in {'1', 'true', 'si', 'yes'}

        try:
            workbook = load_workbook(filename=archivo, data_only=True, read_only=True)
            sheet = workbook[workbook.sheetnames[0]]
        except Exception as exc:
            return Response({'detail': f'No se pudo leer el archivo: {exc}'}, status=400)

        iterator = sheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if not header_row:
            return Response({'detail': 'El archivo no contiene encabezados.'}, status=400)

        header_map = {self._normalize_header(value): idx for idx, value in enumerate(header_row) if value is not None}

        def col_idx(candidates):
            for candidate in candidates:
                idx = header_map.get(candidate)
                if idx is not None:
                    return idx
            return None

        col_detalle = col_idx(['detalle', 'descripcion'])
        col_partida = col_idx(['partida', 'partida_codigo', 'codigo_partida'])
        col_unidad = col_idx(['unidad', 'unidad_medida', 'uom'])

        missing = []
        if col_detalle is None:
            missing.append('DETALLE/descripcion')
        if col_partida is None:
            missing.append('partida/partida_codigo')
        if missing:
            return Response(
                {
                    'detail': 'Faltan columnas requeridas en el Excel.',
                    'faltantes': missing,
                    'encabezados_detectados': list(header_map.keys()),
                },
                status=400,
            )

        stats = {
            'filas_procesadas': 0,
            'filas_vacias': 0,
            'filas_invalidas': 0,
            'items_nuevos_detectados': 0,
            'items_creados': 0,
            'omitidos_total': 0,
            'omitidos_sin_detalle': 0,
            'omitidos_sin_partida': 0,
            'omitidos_duplicado_en_excel': 0,
            'omitidos_duplicado_en_bd': 0,
            'detalle_truncado': 0,
            'dry_run': dry_run,
            'partidas_unicas_detectadas': 0,
        }
        omitidos = []

        partidas_detectadas = set()
        to_create = []
        seen_excel_keys = set()
        existing_keys = set()

        for item in ItemCatalogo.objects.all().only('detalle', 'partida'):
            detail_key = str(item.detalle or '').strip().lower()
            partida_key = str(item.partida or '').strip()
            if detail_key and partida_key:
                existing_keys.add((detail_key, partida_key))

        def get_cell(row, idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for excel_row_num, row in enumerate(iterator, start=2):
            detalle = str(get_cell(row, col_detalle) or '').strip()
            partida_codigo = self._normalize_code(get_cell(row, col_partida))
            unidad_medida = str(get_cell(row, col_unidad) or 'Sin unidad').strip() if col_unidad is not None else 'Sin unidad'

            if not detalle and not partida_codigo:
                stats['filas_vacias'] += 1
                continue

            stats['filas_procesadas'] += 1
            if not detalle:
                stats['filas_invalidas'] += 1
                stats['omitidos_total'] += 1
                stats['omitidos_sin_detalle'] += 1
                if len(omitidos) < 200:
                    omitidos.append(
                        {
                            'fila': excel_row_num,
                            'detalle': detalle,
                            'partida': partida_codigo,
                            'motivo': 'Detalle vacio: la fila se omite automaticamente.',
                        }
                    )
                continue

            if not partida_codigo:
                stats['filas_invalidas'] += 1
                stats['omitidos_total'] += 1
                stats['omitidos_sin_partida'] += 1
                if len(omitidos) < 200:
                    omitidos.append(
                        {
                            'fila': excel_row_num,
                            'detalle': detalle,
                            'partida': partida_codigo,
                            'motivo': 'Partida presupuestaria vacia: la fila se omite automaticamente.',
                        }
                    )
                continue

            if len(detalle) > 255:
                detalle = detalle[:255]
                stats['detalle_truncado'] += 1

            partidas_detectadas.add(partida_codigo)
            detail_key = detalle.lower()
            composite_key = (detail_key, partida_codigo)

            if dry_run:
                if composite_key in seen_excel_keys:
                    stats['omitidos_total'] += 1
                    stats['omitidos_duplicado_en_excel'] += 1
                    if len(omitidos) < 200:
                        omitidos.append(
                            {
                                'fila': excel_row_num,
                                'detalle': detalle,
                                'partida': partida_codigo,
                                'motivo': 'Duplicado dentro del mismo archivo Excel.',
                            }
                        )
                    continue

                if composite_key in existing_keys:
                    stats['omitidos_total'] += 1
                    stats['omitidos_duplicado_en_bd'] += 1
                    if len(omitidos) < 200:
                        omitidos.append(
                            {
                                'fila': excel_row_num,
                                'detalle': detalle,
                                'partida': partida_codigo,
                                'motivo': 'Ya existe en la base de datos (mismo detalle y partida).',
                            }
                        )
                    continue

                seen_excel_keys.add(composite_key)
                stats['items_nuevos_detectados'] += 1
                continue

            if composite_key in seen_excel_keys:
                stats['omitidos_total'] += 1
                stats['omitidos_duplicado_en_excel'] += 1
                if len(omitidos) < 200:
                    omitidos.append(
                        {
                            'fila': excel_row_num,
                            'detalle': detalle,
                            'partida': partida_codigo,
                            'motivo': 'Duplicado dentro del mismo archivo Excel.',
                        }
                    )
                continue

            if composite_key in existing_keys:
                stats['omitidos_total'] += 1
                stats['omitidos_duplicado_en_bd'] += 1
                if len(omitidos) < 200:
                    omitidos.append(
                        {
                            'fila': excel_row_num,
                            'detalle': detalle,
                            'partida': partida_codigo,
                            'motivo': 'Ya existe en la base de datos (mismo detalle y partida).',
                        }
                    )
                continue

            to_create.append(
                ItemCatalogo(
                    partida=partida_codigo,
                    detalle=detalle,
                    unidad_medida=unidad_medida or 'Sin unidad',
                )
            )
            seen_excel_keys.add(composite_key)
            stats['items_nuevos_detectados'] += 1

        if not dry_run:
            with transaction.atomic():
                if to_create:
                    ItemCatalogo.objects.bulk_create(to_create, batch_size=1000)
                    stats['items_creados'] = len(to_create)
        stats['partidas_unicas_detectadas'] = len(partidas_detectadas)

        mensaje = (
            'Importacion completada. Solo se cargaron registros nuevos. '
            'Los registros omitidos deben revisarse e ingresarse manualmente si corresponde.'
        )

        return Response(
            {
                'detail': mensaje,
                'resumen': stats,
                'omitidos': omitidos,
            },
            status=200,
        )

    def get_queryset(self):
        qs = ItemCatalogo.objects.all()
        partida_codigo = (
            self.request.query_params.get('partida')
            or self.request.query_params.get('partida_codigo')
            or self.request.query_params.get('partida_id')
        )
        if partida_codigo:
            return qs.filter(partida=str(partida_codigo).strip())
        # No devolver listado global; exigir filtro por partida
        return qs

    def list(self, request, *args, **kwargs):
        partida_codigo = request.query_params.get('partida') or request.query_params.get('partida_codigo') or request.query_params.get('partida_id')
        if not partida_codigo:
            return Response({'detail': "El parámetro 'partida' es obligatorio para listar items."}, status=400)
        return super().list(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        # Mantener compatibilidad: si llega partida_id usarlo como codigo de partida
        if 'partida' not in data and 'partida_id' in data:
            data['partida'] = str(data.get('partida_id')).strip()
        if 'partida' not in data:
            return Response({'detail': "El campo 'partida' es obligatorio para crear un item."}, status=400)

        detalle = str(data.get('detalle') or '').strip()
        replace_existing = str(data.get('replace_existing', 'false')).strip().lower() in {'1', 'true', 'si', 'yes'}
        if detalle:
            existente = ItemCatalogo.objects.filter(detalle__iexact=detalle).first()
            if existente and not replace_existing:
                return Response(
                    {
                        'detail': 'Ya existe un item con el mismo DETALLE.',
                        'requires_confirmation': True,
                        'duplicate_item_id': existente.id,
                        'duplicate_detalle': existente.detalle,
                    },
                    status=409,
                )
            if existente and replace_existing:
                serializer = self.get_serializer(existente, data=data, partial=True)
                serializer.is_valid(raise_exception=True)
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)



class OperacionCatalogoViewSet(viewsets.ModelViewSet):
    serializer_class = OperacionCatalogoSerializer
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    _TABLE_SETTINGS = {
        'vertical_strategy': 'lines',
        'horizontal_strategy': 'lines',
        'snap_tolerance': 4,
        'join_tolerance': 4,
        'edge_min_length': 3,
        'min_words_vertical': 1,
        'min_words_horizontal': 1,
        'intersection_tolerance': 5,
    }

    @staticmethod
    def _normalize_text(value):
        if value is None:
            return ''
        text = str(value).replace('\xa0', ' ')
        text = unicodedata.normalize('NFKC', text)
        text = re.sub(r'\s+', ' ', text)
        return text.strip()

    @classmethod
    def _normalize_key(cls, value):
        return cls._normalize_text(value).lower()

    @classmethod
    def _clean_cell_text(cls, value):
        text = cls._normalize_text(value)
        if not text:
            return ''
        text = text.replace('•', '\n')
        text = text.replace('–', '-')
        parts = [part.strip(' -') for part in re.split(r'[\n\r]+', text) if part.strip(' -')]
        return '\n'.join(parts)

    @classmethod
    def _is_header_row(cls, row):
        text = ' '.join(cls._normalize_key(cell) for cell in row if cls._normalize_key(cell))
        headers = ['servicios', 'procesos', 'operaciones', 'productos', 'intermedios', 'indicador']
        hits = sum(1 for token in headers if token in text)
        return hits >= 3

    @classmethod
    def _resolve_column_map(cls, header_row):
        def norm(cell):
            return cls._normalize_key(cell)

        headers = [norm(cell) for cell in (header_row or [])]

        def find_index(candidates):
            for idx, header in enumerate(headers):
                if not header:
                    continue
                for candidate in candidates:
                    if candidate in header:
                        return idx
            return None

        servicio_idx = find_index(['servicio', 'servicios'])
        proceso_idx = find_index(['proceso', 'procesos'])
        operacion_idx = find_index(['operacion', 'operaciones'])
        producto_idx = find_index(['producto intermedio', 'productos intermedios', 'producto terminal', 'productos terminales', 'producto', 'productos'])
        indicador_idx = find_index(['indicador', 'indicadores'])

        # Si no se detecta encabezado por texto, usa posición por defecto.
        if operacion_idx is None:
            operacion_idx = 2 if len(headers) >= 3 else 0
        if proceso_idx is None:
            proceso_idx = 1 if len(headers) >= 2 else 0
        if producto_idx is None:
            producto_idx = 3 if len(headers) >= 4 else None
        if indicador_idx is None:
            indicador_idx = 4 if len(headers) >= 5 else (3 if len(headers) >= 4 else None)

        return {
            'servicio': servicio_idx,
            'proceso': proceso_idx,
            'operacion': operacion_idx,
            'producto_intermedio': producto_idx,
            'indicador': indicador_idx,
        }

    @classmethod
    def _get_by_index(cls, row, idx):
        if idx is None:
            return ''
        if idx < 0 or idx >= len(row):
            return ''
        return cls._clean_cell_text(row[idx])

    @classmethod
    def _extract_heading_for_table(cls, page, table_top, previous_heading=''):
        try:
            chars = page.chars
        except Exception:
            return previous_heading

        lines = {}
        for char in chars:
            if char.get('bottom', 0) > table_top - 2:
                continue
            line_key = round(float(char.get('top', 0)), 1)
            lines.setdefault(line_key, []).append(char)

        if not lines:
            return previous_heading

        header_terms = {
            'servicios',
            'procesos',
            'operaciones',
            'productos',
            'productos intermedios',
            'indicador',
            'intermedios',
        }
        selected_lines = []
        started = False
        for line_top, line_words in sorted(lines.items(), reverse=True):
            line_words = sorted(line_words, key=lambda item: item.get('x0', 0))
            text = cls._normalize_text(''.join(word.get('text', '') for word in line_words))
            if not text:
                continue
            low = text.lower()
            if low in header_terms:
                if started:
                    break
                continue
            if re.fullmatch(r'\d+', low):
                if started:
                    break
                continue
            avg_size = 0
            if line_words:
                avg_size = sum(float(word.get('size') or 0) for word in line_words) / len(line_words)
            if avg_size < 10.5:
                if started:
                    break
                continue

            selected_lines.append(text)
            started = True

        if not selected_lines:
            return previous_heading

        selected_lines.reverse()
        heading = cls._normalize_text(' '.join(selected_lines))
        return heading or previous_heading

    @classmethod
    def _split_table_row(cls, row):
        values = [cls._clean_cell_text(cell) for cell in row]
        while len(values) < 5:
            values.append('')
        return values[:5]

    @classmethod
    def _split_multivalue_field(cls, value):
        text = cls._clean_cell_text(value)
        if not text:
            return []
        parts = [part.strip() for part in re.split(r'[\n\r]+', text) if part.strip()]
        return parts or [text]

    @classmethod
    def _split_indicator_values(cls, value):
        text = cls._clean_cell_text(value)
        if not text:
            return []

        # Primero, respetar separaciones explícitas por saltos de línea.
        lines = [part.strip() for part in re.split(r'[\n\r]+', text) if part.strip()]
        if len(lines) > 1:
            return lines

        # Fallback para PDFs donde varios indicadores llegan concatenados en una sola línea.
        marker_pattern = re.compile(r'(?:\bNo\s+de\b|\bN\s*[°ºo�]?\s*de\b)', flags=re.IGNORECASE)
        markers = list(marker_pattern.finditer(text))
        if len(markers) <= 1:
            return [text]

        chunks = []
        for idx, marker in enumerate(markers):
            start = marker.start()
            end = markers[idx + 1].start() if idx + 1 < len(markers) else len(text)
            chunk = text[start:end].strip(' ;,-')
            if chunk:
                chunks.append(chunk)

        return chunks or [text]

    @classmethod
    def _row_has_content(cls, values):
        return any(bool(cls._normalize_text(value)) for value in values)

    def _expand_pdf_row(self, row_data, heading, page_index):
        service = self._clean_cell_text(row_data[0])
        process = self._clean_cell_text(row_data[1])
        operation = self._clean_cell_text(row_data[2])
        product = self._clean_cell_text(row_data[3])
        indicator = self._clean_cell_text(row_data[4])

        indicators = self._split_indicator_values(indicator) or ['']
        if not indicators:
            indicators = ['']

        rows = []
        for indicator_item in indicators:
            rows.append(
                {
                    'direccion': heading or f'Pagina {page_index}',
                    'servicio': service,
                    'proceso': process,
                    'operacion': operation,
                    'producto_intermedio': product,
                    'indicador': self._clean_cell_text(indicator_item),
                }
            )
        return rows

    def _extract_pdf_rows(self, archivo):
        filas = []
        with pdfplumber.open(archivo) as pdf:
            previous_heading = ''
            for page_index, page in enumerate(pdf.pages, start=1):
                tables = page.find_tables(table_settings=self._TABLE_SETTINGS) or []
                if not tables:
                    # Fallback para PDFs donde no se detectan bien las líneas de tabla.
                    extracted_tables = page.extract_tables() or []
                    for raw in extracted_tables:
                        class RawTable:
                            def __init__(self, rows):
                                self._rows = rows
                                self.bbox = (0, 0, 0, 0)

                            def extract(self):
                                return self._rows

                        tables.append(RawTable(raw))

                for table in tables:
                    try:
                        table_rows = table.extract()
                    except Exception:
                        continue

                    if not table_rows:
                        continue

                    heading = self._extract_heading_for_table(page, table.bbox[1], previous_heading)
                    if heading:
                        previous_heading = heading

                    current_service = ''
                    current_process = ''
                    current_operation = ''
                    current_product = ''
                    column_map = {
                        'servicio': 0,
                        'proceso': 1,
                        'operacion': 2,
                        'producto_intermedio': 3,
                        'indicador': 4,
                    }

                    for row_index, row in enumerate(table_rows):
                        normalized_row = self._split_table_row(row)
                        if row_index == 0 and self._is_header_row(normalized_row):
                            column_map = self._resolve_column_map(normalized_row)
                            continue

                        if not self._row_has_content(normalized_row):
                            continue

                        service = self._get_by_index(normalized_row, column_map.get('servicio'))
                        process = self._get_by_index(normalized_row, column_map.get('proceso'))
                        operation = self._get_by_index(normalized_row, column_map.get('operacion'))
                        product = self._get_by_index(normalized_row, column_map.get('producto_intermedio'))
                        indicator = self._get_by_index(normalized_row, column_map.get('indicador'))

                        # Si la tabla no trae columna de servicio (caso 4 columnas), conserva el último servicio.
                        if service:
                            current_service = service
                        else:
                            service = current_service

                        if process:
                            current_process = process
                        else:
                            process = current_process

                        if operation:
                            current_operation = operation
                        else:
                            operation = current_operation

                        if product:
                            current_product = product
                        else:
                            product = current_product

                        merged_row = [service, process, operation, product, indicator]
                        if not self._row_has_content(merged_row):
                            continue

                        filas.extend(self._expand_pdf_row(merged_row, heading or previous_heading, page_index))
        return filas

    @staticmethod
    def _row_key(row):
        return (
            row['direccion'].lower(),
            row['servicio'].lower(),
            row['proceso'].lower(),
            row['operacion'].lower(),
            row['producto_intermedio'].lower(),
            row['indicador'].lower(),
        )

    @action(detail=False, methods=['get', 'post'], url_path='importar-pdf')
    def importar_pdf(self, request, *args, **kwargs):
        if request.method == 'GET':
            return Response(
                {
                    'detail': 'Endpoint de importación de catálogo de indicadores desde PDF.',
                    'uso': 'Envíe POST multipart/form-data con el archivo PDF en el campo "archivo".',
                },
                status=200,
            )

        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'detail': 'Debe adjuntar un archivo PDF en el campo "archivo".'}, status=400)

        dry_run = str(request.data.get('dry_run', 'false')).strip().lower() in {'1', 'true', 'si', 'yes'}

        try:
            filas = self._extract_pdf_rows(archivo)
        except Exception as exc:
            return Response({'detail': f'No se pudo leer el PDF: {exc}'}, status=400)

        if not filas:
            return Response(
                {
                    'detail': 'No se detectaron tablas o filas válidas en el PDF.',
                    'sugerencia': 'Verifique que el archivo sea un PDF con texto y tablas legibles.',
                },
                status=400,
            )

        direction_names = []
        seen_direction_names = set()
        for row in filas:
            direccion_nombre = self._normalize_text(row.get('direccion'))
            if not direccion_nombre:
                continue
            key = self._normalize_key(direccion_nombre)
            if key in seen_direction_names:
                continue
            seen_direction_names.add(key)
            direction_names.append(direccion_nombre)

        stats = {
            'filas_detectadas': len(filas),
            'filas_procesadas': 0,
            'filas_vacias': 0,
            'filas_invalidas': 0,
            'direcciones_detectadas': 0,
            'direcciones_creadas': 0,
            'direcciones_reutilizadas': 0,
            'operaciones_creadas': 0,
            'operaciones_actualizadas': 0,
            'registros_expandidos': len(filas),
            'omitidos_total': 0,
            'omitidos_sin_direccion': 0,
            'omitidos_sin_operacion': 0,
            'omitidos_sin_indicador': 0,
            'omitidos_indicador_duplicado': 0,
            'dry_run': dry_run,
        }
        omitidos = []
        direccion_cache = {}
        valid_rows = []
        seen_indicator_keys = set()
        existing_indicator_map = {}

        for op in OperacionCatalogo.objects.select_related('direccion').only(
            'id',
            'indicador',
            'direccion__nombre',
            'servicio',
            'proceso',
            'operacion',
            'producto_intermedio',
            'direccion',
        ):
            indicador_text = self._normalize_text(op.indicador)
            if not indicador_text:
                continue
            key = (self._normalize_key(op.direccion.nombre), self._normalize_key(indicador_text))
            if key not in existing_indicator_map:
                existing_indicator_map[key] = op

        if not dry_run:
            for direccion_nombre in direction_names:
                direccion_obj, created = Direccion.objects.get_or_create(nombre=direccion_nombre)
                direccion_cache[direccion_nombre] = direccion_obj
                if created:
                    stats['direcciones_creadas'] += 1
                else:
                    stats['direcciones_reutilizadas'] += 1
        else:
            for direccion_nombre in direction_names:
                direccion_cache[direccion_nombre] = direccion_nombre

        for fila_num, row in enumerate(filas, start=1):
            direccion_nombre = self._normalize_text(row.get('direccion'))
            servicio = self._normalize_text(row.get('servicio'))
            proceso = self._normalize_text(row.get('proceso'))
            operacion = self._normalize_text(row.get('operacion'))
            producto_intermedio = self._normalize_text(row.get('producto_intermedio'))
            indicador = self._normalize_text(row.get('indicador'))

            if not direccion_nombre:
                stats['filas_invalidas'] += 1
                stats['omitidos_total'] += 1
                stats['omitidos_sin_direccion'] += 1
                if len(omitidos) < 200:
                    omitidos.append({
                        'fila': fila_num,
                        'motivo': 'No se pudo identificar la dirección del bloque.',
                    })
                continue

            if not operacion:
                stats['filas_invalidas'] += 1
                stats['omitidos_total'] += 1
                stats['omitidos_sin_operacion'] += 1
                if len(omitidos) < 200:
                    omitidos.append({
                        'fila': fila_num,
                        'direccion': direccion_nombre,
                        'motivo': 'La fila no tiene operación y no se importó.',
                    })
                continue

            if not indicador:
                stats['filas_invalidas'] += 1
                stats['omitidos_total'] += 1
                stats['omitidos_sin_indicador'] += 1
                if len(omitidos) < 200:
                    omitidos.append({
                        'fila': fila_num,
                        'direccion': direccion_nombre,
                        'operacion': operacion,
                        'motivo': 'La fila no tiene indicador y no se importó.',
                    })
                continue

            if indicador:
                indicator_key = (self._normalize_key(direccion_nombre), self._normalize_key(indicador))
                if indicator_key in seen_indicator_keys:
                    stats['filas_invalidas'] += 1
                    stats['omitidos_total'] += 1
                    stats['omitidos_indicador_duplicado'] += 1
                    if len(omitidos) < 200:
                        omitidos.append({
                            'fila': fila_num,
                            'direccion': direccion_nombre,
                            'indicador': indicador,
                            'motivo': 'Indicador duplicado para la misma dirección.',
                        })
                    continue
                seen_indicator_keys.add(indicator_key)

            stats['filas_procesadas'] += 1
            valid_rows.append(
                {
                    'direccion': direccion_nombre,
                    'servicio': servicio,
                    'proceso': proceso,
                    'operacion': operacion,
                    'producto_intermedio': producto_intermedio,
                    'indicador': indicador,
                }
            )

        if dry_run:
            dry_creadas = 0
            dry_actualizadas = 0
            for row in valid_rows:
                indicator_key = (
                    self._normalize_key(row['direccion']),
                    self._normalize_key(row['indicador']),
                )
                if indicator_key in existing_indicator_map:
                    dry_actualizadas += 1
                else:
                    dry_creadas += 1

            stats['operaciones_creadas'] = dry_creadas
            stats['operaciones_actualizadas'] = dry_actualizadas
            stats['direcciones_detectadas'] = len(direction_names)
            return Response(
                {
                    'detail': 'Vista previa de importación generada.',
                    'resumen': stats,
                    'omitidos': omitidos,
                },
                status=200,
            )

        to_create = []
        to_update = []
        for row in valid_rows:
            direccion_nombre = row['direccion']
            servicio = row['servicio']
            proceso = row['proceso']
            operacion = row['operacion']
            producto_intermedio = row['producto_intermedio']
            indicador = row['indicador']

            direccion_obj = direccion_cache.get(direccion_nombre)
            if direccion_obj is None:
                direccion_obj, created = Direccion.objects.get_or_create(nombre=direccion_nombre)
                direccion_cache[direccion_nombre] = direccion_obj
                if created:
                    stats['direcciones_creadas'] += 1
                else:
                    stats['direcciones_reutilizadas'] += 1

            indicator_key = (self._normalize_key(direccion_nombre), self._normalize_key(indicador))
            existing_op = existing_indicator_map.get(indicator_key)
            if existing_op is not None:
                existing_op.direccion = direccion_obj
                existing_op.servicio = servicio
                existing_op.proceso = proceso
                existing_op.operacion = operacion
                existing_op.producto_intermedio = producto_intermedio
                existing_op.indicador = indicador
                to_update.append(existing_op)
                continue

            to_create.append(
                OperacionCatalogo(
                    direccion=direccion_obj,
                    servicio=servicio,
                    proceso=proceso,
                    operacion=operacion,
                    producto_intermedio=producto_intermedio,
                    indicador=indicador,
                )
            )

        with transaction.atomic():
            if to_update:
                OperacionCatalogo.objects.bulk_update(
                    to_update,
                    ['direccion', 'servicio', 'proceso', 'operacion', 'producto_intermedio', 'indicador'],
                )
            if to_create:
                OperacionCatalogo.objects.bulk_create(to_create, batch_size=1000)

        stats['operaciones_actualizadas'] = len(to_update)
        stats['operaciones_creadas'] = len(to_create)
        stats['direcciones_detectadas'] = len(direction_names)

        return Response(
            {
                'detail': 'Importación de indicadores completada.',
                'resumen': stats,
                'omitidos': omitidos,
            },
            status=200,
        )

    def get_queryset(self):
        qs = OperacionCatalogo.objects.select_related('direccion').all()

        direccion_id = self.request.query_params.get('direccion_id')
        if direccion_id:
            try:
                qs = qs.filter(direccion_id=int(direccion_id))
            except (ValueError, TypeError):
                return qs.none()

        servicio = (self.request.query_params.get('servicio') or '').strip()
        proceso = (self.request.query_params.get('proceso') or '').strip()
        operacion = (self.request.query_params.get('operacion') or '').strip()
        indicador = (self.request.query_params.get('indicador') or '').strip()
        direccion_nombre = (self.request.query_params.get('direccion') or '').strip()
        q = (self.request.query_params.get('q') or self.request.query_params.get('search') or '').strip()

        if servicio:
            qs = qs.filter(servicio__icontains=servicio)
        if proceso:
            qs = qs.filter(proceso__icontains=proceso)
        if operacion:
            qs = qs.filter(operacion__icontains=operacion)
        if indicador:
            qs = qs.filter(indicador__icontains=indicador)
        if direccion_nombre:
            qs = qs.filter(direccion__nombre__icontains=direccion_nombre)
        if q:
            qs = qs.filter(
                Q(direccion__nombre__icontains=q)
                | Q(servicio__icontains=q)
                | Q(proceso__icontains=q)
                | Q(operacion__icontains=q)
                | Q(producto_intermedio__icontains=q)
                | Q(indicador__icontains=q)
            )

        return qs

    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        if 'direccion_id' not in request.data and 'direccion' not in request.data:
            return Response({'detail': "El campo 'direccion_id' es obligatorio para crear una operación."}, status=400)

        direccion_id = request.data.get('direccion_id') or request.data.get('direccion')
        indicador = self._normalize_text(request.data.get('indicador'))
        if direccion_id and indicador:
            try:
                direccion_id = int(direccion_id)
            except (ValueError, TypeError):
                return Response({'detail': "El campo 'direccion_id' debe ser un entero válido."}, status=400)

            duplicated = OperacionCatalogo.objects.filter(direccion_id=direccion_id, indicador__iexact=indicador).exists()
            if duplicated:
                return Response({'detail': 'El indicador ya existe para esta dirección.'}, status=409)

        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()

        direccion_id = request.data.get('direccion_id') or request.data.get('direccion') or instance.direccion_id
        indicador = self._normalize_text(request.data.get('indicador')) if 'indicador' in request.data else self._normalize_text(instance.indicador)

        if indicador:
            try:
                direccion_id = int(direccion_id)
            except (ValueError, TypeError):
                return Response({'detail': "El campo 'direccion_id' debe ser un entero válido."}, status=400)

            duplicated = OperacionCatalogo.objects.filter(
                direccion_id=direccion_id,
                indicador__iexact=indicador,
            ).exclude(id=instance.id).exists()
            if duplicated:
                return Response({'detail': 'El indicador ya existe para esta dirección.'}, status=409)

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return Response(serializer.data)


class DireccionViewSet(viewsets.ModelViewSet):
    queryset = Direccion.objects.all()
    serializer_class = DireccionSerializer
    permission_classes = [AllowAny]


class PartidaPresupuestariaViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = PartidaCatalogoSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        partidas = (
            ItemCatalogo.objects
            .exclude(partida='')
            .values_list('partida', flat=True)
            .distinct()
            .order_by('partida')
        )
        return [{'id': p, 'codigo': p, 'nombre': p} for p in partidas]


class ItemCatalogoReadOnlyViewSet(viewsets.ReadOnlyModelViewSet):
    """Endpoint solo-lectura para listar/ver detalles de items junto con su partida.

    Este viewset expone GET / y GET /<id>/ sin exigir el filtro por partida_id.
    """
    queryset = ItemCatalogo.objects.all()
    serializer_class = ItemCatalogoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        partida_codigo = (
            self.request.query_params.get('partida')
            or self.request.query_params.get('partida_codigo')
            or self.request.query_params.get('partida_id')
        )
        if partida_codigo:
            qs = qs.filter(partida=str(partida_codigo).strip())

        search = (self.request.query_params.get('q') or self.request.query_params.get('search') or '').strip()
        if search:
            qs = qs.filter(
                Q(detalle__icontains=search)
                | Q(unidad_medida__icontains=search)
                | Q(partida__icontains=search)
            )

        return qs.order_by('id')

    def list(self, request, *args, **kwargs):
        search = (request.query_params.get('q') or request.query_params.get('search') or '').strip()
        queryset = self.filter_queryset(self.get_queryset())

        # En búsquedas devolver una muestra mayor sin paginación para UX más fluida.
        if search:
            serializer = self.get_serializer(queryset[:200], many=True)
            return Response(serializer.data)

        return super().list(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset()).order_by('id')
        wb = Workbook()
        ws = wb.active
        ws.title = 'CatalogoItems'

        headers = ['DETALLE', 'partida', 'UNIDAD_MEDIDA']
        ws.append(headers)

        header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_side = Side(style='thin', color='D9D9D9')
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        for it in queryset.iterator():
            ws.append([
                str(it.detalle or ''),
                str(it.partida or ''),
                str(it.unidad_medida or ''),
            ])

        # Estilo del cuerpo y mejor legibilidad
        body_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = body_alignment
                cell.border = border

        # Congelar encabezado y habilitar autofiltro
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

        # Autoajustar ancho de columnas según contenido
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                text = str(value)
                if len(text) > max_length:
                    max_length = len(text)
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 14), 90)

        ws.row_dimensions[1].height = 24

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="catalogo_items.xlsx"'
        wb.save(response)
        return response


class OperacionCatalogoReadOnlyViewSet(viewsets.ReadOnlyModelViewSet):
    """Exponer todas las operaciones sin requerir filtro por dirección."""
    queryset = OperacionCatalogo.objects.select_related('direccion').all()
    serializer_class = OperacionCatalogoSerializer
    permission_classes = [AllowAny]


